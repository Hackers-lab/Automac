import pandas as pd
from tkinter import Tk, Label, Button, OptionMenu, StringVar, Entry, messagebox, ttk, filedialog, Menu, Toplevel, Frame, END
import tkinter as tk
import datetime as dt
from tkinter.filedialog import askopenfilename, askdirectory
import csv
import socket, random
import os
import subprocess
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import webbrowser

dc_file_path = ""
zone_file_path = ""
output_dir = ""
dc_df = pd.DataFrame()
zone_df = pd.DataFrame()
filtered_dc_df_list = []


def apply_filters_callback(event):
    apply_filters()  # Call your apply_filters function here
def browse_dc_file():
    global dc_file_path, dc_file_path, zone_file_path, dc_df, zone_df, output_dir
    dc_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    settings_file_path = "c:/dcrc/settings.csv"

    try:
        with open(settings_file_path, "r") as file:
            reader = csv.reader(file)
            rows = list(reader)
            if len(rows) >= 2 and len(rows[1]) >= 2:
                zone_file_path = rows[1][1]
            else:
                raise ValueError("Invalid settings.csv file format.")
    except IOError:
        zone_file_path = ""
        #messagebox.showerror("Error", "Failed to read settings.csv.")
    except ValueError as e:
        zone_file_path = ""
        messagebox.showerror("Error", str(e))
    if not os.path.isfile(zone_file_path):
        zone_file_path = ""
        messagebox.showerror("Zone file not found", "Please update the zone path from settings menu first.")

    try:
        with open(settings_file_path, "r") as file:
            reader = csv.reader(file)
            rows = list(reader)
            if len(rows) >= 2 and len(rows[1]) >= 2:
                try:
                    output_dir = rows[0][1]
                except IndexError:
                    messagebox.showerror("Error", "Output directory is not set. Please set it from the settings menu.")
                    return
            else:
                raise ValueError("Invalid output directory path in settings.csv.")
    except IOError:
        output_dir = ""
        #messagebox.showerror("Error", "Failed to read settings.csv.")
        return
    except ValueError as e:
        output_dir = ""
        messagebox.showerror("Error", str(e))
        return



    if dc_file_path and zone_file_path:
        # Read the Excel files
        dc_df = pd.read_excel(dc_file_path)

        # Verify column names in the DataFrame
        required_columns = ['MRU', 'Base Class', 'Discon Status', 'D2 Net O/S']
        missing_columns = [col for col in required_columns if col not in dc_df.columns]

        if missing_columns:
            messagebox.showerror("Error", f"Please verify the selected disconnection file.")
            return   

        zone_df = pd.read_excel(zone_file_path)
        if dc_file_path:
            dcc_label.config(text=dc_file_path,bg="white", fg="green")

        # Show dropdown list of agencies
        agencies = zone_df.columns.tolist()  # Assuming each agency name is in a separate column
        agencies.insert(0, "ALL")

        # Prompt user for agency, amount, and filter value using Tkinter GUI
        agency_label = Label(left_frame, text="Select an agency:")
        agency_label.grid(row=3, column=0,pady=5, sticky="nw")

        # Create a style using ttk
        style = ttk.Style(root)

        # Configure the style for the OptionMenu
        style.configure('Custom.TMenubutton', foreground='black')




        agency_var.set(agencies[0])  # Set default value
        agency_option_menu = ttk.OptionMenu(left_frame, agency_var, agencies[0], *agencies, style="Custom.TMenubutton", command=apply_filters_callback)
        agency_option_menu.grid(row=3, column=1,pady=5, sticky="nw")

        amount_label = Label(left_frame, text="Enter the amount:")
        amount_label.grid(row=5, column=0,pady=5, sticky="nw")

        global amount_entry  # Update to be a global variable
        amount_entry = Entry(left_frame, width=30)
        amount_entry.grid(row=5, column=1,pady=5, sticky="nw")
        amount_entry.bind("<KeyRelease>", apply_filters_callback)

        filter_label = Label(left_frame, text="Select Class:")
        filter_label.grid(row=4, column=0,pady=5, sticky="nw")


        filter_options = ['D', 'C', 'A', 'I']  # Update with your desired filter options
        filter_var.set(filter_options[0])  # Set default value
        filter_option_menu = ttk.OptionMenu(left_frame, filter_var, filter_options[0], *filter_options, style="Custom.TMenubutton", command=apply_filters_callback)
        filter_option_menu.grid(row=4, column=1, pady=5, sticky="nw")
    


        save_as_button.grid(row=6, column=0, padx=10, pady=10, sticky="nw")
        save_label = Label(left_frame, text=f"*Files will be saved at: {output_dir}", bg="white", fg="blue")
        save_label.grid(row=6, column=0, columnspan=2,padx=60, pady=15, sticky="nw")
def apply_filters():
    selected_agency = agency_var.get()
    amount = amount_entry.get()     
    filter_value = filter_var.get()
    global output_dir

    try:
        amount = float(amount)
    except ValueError:
        amount = 0
        amount_entry.delete(0, END)  # Clear the current value
        amount_entry.insert(0, "0")  # Set the new value
        #messagebox.showerror("Error", "Invalid amount. Please enter a numeric value.")
        

    global filtered_dc_df_list
    filtered_dc_df_list = []

    if selected_agency == "ALL":
        for agency in zone_df.columns:
            if agency != "ALL":
                filtered_dc_df = dc_df[(dc_df['MRU'].isin(zone_df[agency])) &
                                       (dc_df['Base Class'] == filter_value) &
                                       (dc_df['Discon Status'].isnull()) &
                                       (dc_df['D2 Net O/S'] >= amount)]
                filtered_dc_df = filtered_dc_df.drop(columns=["off_code", "Class", "Nature of Conn", "Gov/Non-Gov", "Discon Status", "Discon Date", "Gis Pole"])
                filtered_dc_df_list.append((filtered_dc_df, agency))
    else:
        filtered_dc_df = dc_df[(dc_df['MRU'].isin(zone_df[selected_agency])) &
                               (dc_df['Base Class'] == filter_value) &
                               (dc_df['Discon Status'].isnull()) &
                               (dc_df['D2 Net O/S'] > amount)]
        filtered_dc_df = filtered_dc_df.drop(columns=["off_code", "Class", "Nature of Conn", "Gov/Non-Gov", "Discon Status", "Discon Date", "Gis Pole"])
        filtered_dc_df_list = [(filtered_dc_df, selected_agency)]

    if "right_frame" in globals() and right_frame:
        right_frame.destroy()
    right_frame = tk.Frame(root, bg="white")
    right_frame.grid(row=0, column=1, rowspan=7, padx=10, pady=10, sticky="nsew")
    root.columnconfigure(1, weight=1)
    root.rowconfigure(0, weight=1)

    # Create the treeview widget
    treeview = ttk.Treeview(right_frame)
    treeview.grid(row=0, column=0, sticky="nsew")

    # Create a vertical scrollbar for the treeview
    y_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=treeview.yview)
    y_scrollbar.grid(row=0, column=1, sticky="ns")

    # Create a horizontal scrollbar for the treeview
    x_scrollbar = ttk.Scrollbar(right_frame, orient="horizontal", command=treeview.xview)
    x_scrollbar.grid(row=1, column=0, sticky="ew")

    # Configure treeview columns
    columns = filtered_dc_df.columns.tolist()
    treeview["columns"] = columns
    treeview.column("#0", width=0, stretch="NO")  # Hide the first empty column

    for column in columns:
        treeview.heading(column, text=column)
        treeview.column(column, width=100)

    # Insert data into treeview
    for filtered_df, agency in filtered_dc_df_list:
        for index, row in filtered_df.iterrows():
            treeview.insert("", "end", text=agency, values=row.tolist())

    # Configure treeview scrollbars
    treeview.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
    y_scrollbar.configure(command=treeview.yview)
    x_scrollbar.configure(command=treeview.xview)

    # Configure grid weights for right_frame
    right_frame.grid_rowconfigure(0, weight=1)
    right_frame.grid_columnconfigure(0, weight=1)

    # Expand the treeview to fill the available space in right_frame
    treeview.grid(row=0, column=0, sticky="nsew")
def search():
    query = search_entry.get().lower()
    if query:
        right_window = tk.Toplevel(root)
        right_window.title("Search rights")

        for filtered_df, agency in filtered_dc_df_list:
            search_rights = filtered_df[(filtered_df.astype(str).apply(lambda x: x.str.contains(query, case=False)).any(axis=1))]
            if not search_rights.empty:
                search_rights = search_rights[["Agency Name", "Date"]]
                search_rights.insert(0, "Agency", agency)
                search_rights_label = Label(right_window, text=f"Search rights in {agency}:")
                search_rights_label.pack()
                search_rights_table = ttk.Treeview(right_window)
                search_rights_table.pack()

                search_rights_table["columns"] = search_rights.columns.tolist()
                search_rights_table.column("#0", width=0, stretch="NO")

                for column in search_rights.columns:
                    search_rights_table.heading(column, text=column)

                for index, row in search_rights.iterrows():
                    search_rights_table.insert("", "end", values=row.tolist())
def save_as():
    today = dt.date.today().strftime("%Y-%m-%d")
    settings_file_path = "c:/dcrc/settings.csv"
    filter_value = filter_var.get()
    amount = amount_entry.get() 
    global output_dir
    # Validate amount entry
    if not amount.strip():  # Check if the amount string is empty or contains only whitespace
        messagebox.showerror("Error", "Please enter a valid amount.")
        return
    
    

    # Check if output_dir is empty or invalid
    if not output_dir or not os.path.isdir(output_dir):
        output_dir = os.path.join(os.getcwd(), output_dir)

    try:
        os.makedirs(output_dir, exist_ok=True)
    except OSError:
        output_dir = ""
        messagebox.showerror("Error", f"Invalid output directory path in settings.csv: {output_dir}. Please update it from the settings.")
        return

    if output_dir:
        creation_history_file = os.path.join(output_dir, "creation_history.csv")
        creation_history_exists = os.path.exists(creation_history_file)

        with open(creation_history_file, mode='a', newline='') as file:
            writer = csv.writer(file)

            if not creation_history_exists:
                writer.writerow(['Agency Name', 'File Name', 'Date', 'Time', 'Number of Rows', 'Total Value', 'Amount', 'IP Address'])

            # Check if the master file exists
            master_file_path = os.path.join(output_dir, "master.xlsx")
            if os.path.exists(master_file_path):
                # Load the existing master workbook
                master_workbook = openpyxl.load_workbook(master_file_path)
                master_worksheet = master_workbook.active
            else:
                # Create a new master workbook and worksheet
                master_workbook = openpyxl.Workbook()
                master_worksheet = master_workbook.active
                # Write the header row to the master worksheet
                header = filtered_dc_df_list[0][0].columns.tolist()  # Assuming the first DataFrame has the same columns for all agencies
                master_worksheet.append(header)

                # Create a set to store unique rows
            unique_rows = set()

            # Load existing data into the set
            for existing_row in master_worksheet.iter_rows(min_row=2, values_only=True):
                unique_rows.add(tuple(existing_row[:-2]))  # Exclude agency and date columns
            

            for filtered_dc_df, agency in filtered_dc_df_list:
                output_file_path = os.path.join(output_dir, f"{agency}_{filter_value}_{today}.xlsx")
                

                if os.path.exists(output_file_path):
                    overwrite = messagebox.askyesno("File Already Exists", f"The output file for {agency.upper()} already exists. Do you want to overwrite it?")
                    if not overwrite:
                        continue

                filtered_dc_df.to_excel(output_file_path, index=False)

                for index, row in filtered_dc_df.iterrows():
                    values = row.tolist()           
                    values += [agency, today] 
                    #master_worksheet.append(values)
                    if tuple(values[:-2]) not in unique_rows:  # Exclude agency and date columns for comparison
                        unique_rows.add(tuple(values[:-2]))  # Add the row to the set
                        master_worksheet.append(values)

                # Open the workbook
                workbook = openpyxl.load_workbook(output_file_path)
                worksheet = workbook.active

                # Add borders to all cells
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = border


                # Autofit columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)  # Get the column letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                                if max_length > 50:
                                    max_length = 50
                        except TypeError:
                            pass
                    adjusted_width = (max_length)   # Add some padding and adjust the width
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    worksheet.column_dimensions['C'].width = 25
                # Set print settings
                worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
                worksheet.page_margins = PageMargins(top=0.75, left=0.25, right=0.25, bottom=0.25)
                worksheet.page_setup.scale = 80
                worksheet.print_title_cols = 'A:K'
                worksheet.print_title_rows = '1:1'
                worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

                # Add header with the file name
                worksheet.oddHeader.center.text = f"&\"Arial,Bold\"&14{agency}_{filter_value}_{today}"
                worksheet.oddHeader.right.text = "Page-&P/&N"

                # Save the modified workbook
                workbook.save(output_file_path)


                num_rows = len(filtered_dc_df.index)  # Subtract 1 to exclude header row
                total_value = filtered_dc_df['D2 Net O/S'].sum()
                amount = float(amount_entry.get())
                ip_address = socket.gethostbyname(socket.gethostname())
                writer.writerow([agency, f"{agency}_{filter_value}_{today}.xlsx", today, dt.datetime.now().strftime("%H:%M:%S"), num_rows, total_value, amount, ip_address])

            # Save the master workbook
            master_workbook.save(master_file_path)


        messagebox.showinfo("Success", f"Filtered results saved successfully at {output_dir}.")
    else:
        messagebox.showerror("Error", "Invalid output directory path.")
def exit_app():
    if messagebox.askokcancel("Exit", "Do you want to exit the application?"):
        root.quit()
def about_app():
    new_window = Toplevel()
    new_window.title("About DCRC Automac V1.01:")
    new_window.geometry("400x500")

    label_title = Label(new_window, text="DCRC Automac", font=("Helvetica", 16, "bold"))
    label_title.pack(pady=20)

    label_description = Label(new_window, text="The DCRC Automation application is designed to streamline Excel file processing and automate distribution and tracking of disconnection lists. "
                                                "It provides functionalities such as data filtering, file saving, class selection, and agency selection. "
                                                "The application aims to enhance efficiency in managing disconnection reports and improve productivity.", wraplength=380, justify="center")
    label_description.pack(pady=10)

    label_disclaimer = Label(new_window, text="Please note that while the application strives to provide accurate and reliable results, it cannot guarantee the absolute accuracy or completeness of the processed data. "
                                            "Users are advised to review and verify the results obtained from the application.", wraplength=380, justify="center")
    label_disclaimer.pack(pady=10)

    label_contact = Label(new_window, text="For any questions, concerns, or feedback regarding the application, "
                                            "please contact me through je.kushidaccc@gmail.com.", wraplength=380, justify="center")
    label_contact.pack(pady=10)

    label_creator = Label(new_window, text="Created by Pramod Verma")
    label_creator.pack(pady=20)

    label_rights = Label(new_window, text="© All Rights Reserved 2023")
    label_rights.pack(pady=10)

    new_window.mainloop()
def set_zone_with_password():
    ask_password(set_zone)
def set_save_with_password():
    ask_password(set_save)
def ask_password(callback):
    password = dt.date.today().strftime("%Y%m%d")  # Generate password using current date
    password_window = tk.Tk()
    password_window.title("Password")
    password_window.geometry("300x100")  # Set the size of the window
    
    def check_password():
        entered_password = password_entry.get()
        if entered_password == password:
            password_window.destroy()
            callback()
        else:
            random_color = f'#{random.randint(0, 0xFFFFFF):06x}'
            password_label.config(text="Incorrect password. Please try again.", fg=random_color)
    
    password_label = tk.Label(password_window, text="Enter the password to modify:")
    password_label.pack()
    
    
    password_entry = tk.Entry(password_window, show="*")
    password_entry.pack()
    
    submit_button = tk.Button(password_window, text="Submit", command=check_password)
    submit_button.pack()
    
    password_window.mainloop()
def set_zone():
    Tk().withdraw()
    file_path = askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    update_settings_csv(2, 2, file_path)
    messagebox.showinfo("Done", f"The zone file \"{file_path}\" registered!!")
def set_save():
    Tk().withdraw()
    folder_path = askdirectory()
    update_settings_csv(1, 2, folder_path)
    messagebox.showinfo("Done", f"The report/list files will be saved at {folder_path}")
def update_settings_csv(row, column, path):
    file_path = "C:/dcrc/settings.csv"

    # Check if the file exists, create it if it doesn't
    if not os.path.isfile(file_path):
        with open(file_path, "w", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(["Save path"])
            writer.writerow(["Zone file path"])
    # Read the existing data from the file
    rows = []
    with open(file_path, "r") as file:
        reader = csv.reader(file)
        for r in reader:
            rows.append(r)

    # Update the specified cell with the new path
    if row <= len(rows):
        if column <= len(rows[row - 1]):
            rows[row - 1][column - 1] = path
        else:
            rows[row - 1].extend([""] * (column - len(rows[row - 1]) - 1))
            rows[row - 1].append(path)
    else:
        rows.extend([[]] * (row - len(rows) - 1))
        rows.append([""] * (column - 1))
        rows[row - 1].append(path)

    # Write the updated data back to the file
    with open(file_path, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(rows)
def about_ver():
    new_window = Toplevel()
    new_window.title("Version info DCRC Automac V1.01 :")
    new_window.geometry("400x500")

    label_title = Label(new_window, text="Version 1.01", font=("Helvetica", 16, "bold"))
    label_title.pack(pady=20)
    version_title = Label(new_window, text="1.01", font=("Helvetica", 12, "bold"))
    version_title.pack(pady=20)
    label_description = Label(new_window, text="1. User can select zone file using settings menu.\n"
                                                "2. User can select output directory using settings menu.\n"
                                                "3. Data display before saving into excel file.", wraplength=380, justify="left")
    label_description.pack(pady=10)
def open_folder():
    global output_dir
    settings_file_path = "c:/dcrc/settings.csv"
    with open(settings_file_path, "r") as file:
        reader = csv.reader(file)
        rows = list(reader)
        if len(rows) >= 2 and len(rows[1]) >= 2:
            try:
                output_dir = rows[0][1]
            except IndexError:
                messagebox.showerror("Error", "Output directory is not set. Please set it from the settings menu.")
                return    
    webbrowser.open(output_dir)
def open_help_file():
    path = r"C:\dcrc\help\help.html"
    webbrowser.open(path)


# Create the root Tkinter window
root = Tk()
root.title("DCRC Automac V1.01")
root.state('zoomed')

left_frame = tk.Frame(root)
left_frame.grid(row=0, column=0, sticky="nsew")
root.columnconfigure(0, weight=0)
root.rowconfigure(0, weight=1)

dc_label = Label(left_frame, text="Disconnection File PATH:")
dc_label.grid(row=0, column=0,pady=20, sticky="nw")

dcc_label = Label(left_frame, text="Select the DC file from menu.", bg="white", fg="red")
dcc_label.grid(row=0, column=1,padx=5, pady=20, sticky="nw")

agency_var = StringVar(root)
amount_entry = None
filter_var = StringVar(root)

#apply_button = Button(left_frame, text="Apply Filter", command=apply_filters)
save_as_button = Button(left_frame, text="Save", relief="groove", command=save_as)


search_label = Label(left_frame, text="Search:")
search_entry = Entry(left_frame, width=30)
search_button = Button(left_frame, text="Search", command=search)

# Create the menu bar
menu_bar = Menu(root)
root.config(menu=menu_bar)

# Create the File menu
file_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="File", menu=file_menu)

view_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="View", menu=view_menu)

report_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Reports", menu=report_menu)

settings_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Settings", menu=settings_menu)
settings_menu.add_command(label="Set zone file path", command=set_zone_with_password)
settings_menu.add_command(label="Set save file path", command=set_save_with_password)

help_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Help", menu=help_menu)
help_menu.add_command(label="Automac Help", command=open_help_file)

about_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="About", menu=about_menu)
about_menu.add_command(label="About", command=about_app)
about_menu.add_command(label="Version info", command=about_ver)

# Add options to the File menu
file_menu.add_command(label="Open DCRC file", command=browse_dc_file)
file_menu.add_command(label="Save", command=save_as)
file_menu.add_command(label="Output", command=open_folder)
file_menu.add_command(label="Exit", command=exit_app)




root.mainloop()
